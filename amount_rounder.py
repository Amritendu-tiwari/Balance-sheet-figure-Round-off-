from __future__ import annotations
import math
import os
import re
import json
from datetime import datetime
from typing import Optional, Set, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

# ---------------- Utils -----------------
DAYS_HEADER_VARIANTS: Set[str] = {
    "no of days", "no. of days", "number of days", "days", "no days", "no of day"
}

# Headers that imply a year/date column
YEARISH_HEADER_TOKENS: Set[str] = {
    "year", "yr", "fy", "f.y.", "financial year", "fiscal year",
    "ay", "assessment year", "calendar year", "period", "quarter", "qtr"
}

MONTH_TOKENS = {
    "jan","january","feb","february","mar","march","apr","april","may","jun","june",
    "jul","july","aug","august","sep","sept","september","oct","october","nov","november","dec","december"
}

ROW_YEARISH_TOKENS = {
    "as at","as on","year ended","fy","financial year","fiscal year","assessment year","ay","quarter","qtr",
    "balance sheet","profit and loss","statement of"
}

YEAR_REGEXES = [
    re.compile(r"\b(19|20)\d{2}\b"),               # 2024
    re.compile(r"\b(19|20)\d{2}\s*[-/]\s*\d{2}\b") # 2024-25, 2024/25
]

def normalize_header(s: Optional[str]) -> str:
    if s is None:
        return ""
    return " ".join(str(s).replace("\n", " ").strip().lower().split())

def round_half_up_int(x: float) -> int:
    sign = -1 if x < 0 else 1
    a = abs(x)
    return int(sign * math.floor(a + 0.5))

def is_percentage_cell(cell) -> bool:
    fmt = (cell.number_format or "").lower()
    return "%" in fmt

def header_map(ws: Worksheet, header_row: int) -> Dict[int, str]:
    hdrs = {}
    max_col = ws.max_column or 0
    for c in range(1, max_col + 1):
        hdrs[c] = normalize_header(ws.cell(row=header_row, column=c).value)
    return hdrs

# ------------- Core Conversion (division) -------------

def _should_fallback_to_thousand(val: float, lakh_edge_threshold: float) -> bool:
    # If lakhs would be too small/unhelpful, prefer thousands
    return abs(val) < lakh_edge_threshold

def _divide_amount(val: float, mode: str, lakh_edge_threshold: float) -> float:
    """
    Convert by dividing into thousand/lakh with 2 decimal places.
    """
    if mode == "thousand":
        return round(val / 1000.0, 2)
    elif mode == "lakh":
        if _should_fallback_to_thousand(val, lakh_edge_threshold):
            return round(val / 1000.0, 2)
        return round(val / 100000.0, 2)
    elif mode == "auto":
        # If 1 lakh or more -> lakhs; else if moderately large -> thousands; else leave as is
        if abs(val) >= 100000:
            return round(val / 100000.0, 2)
        elif abs(val) >= lakh_edge_threshold:
            return round(val / 1000.0, 2)
        else:
            return val
    else:
        return val

# ---------- Year/date detection ----------

def _cell_has_yearish_text(val: str) -> bool:
    s = normalize_header(val)
    if not s:
        return False
    if any(rx.search(s) for rx in YEAR_REGEXES):
        return True
    if any(tok in s for tok in ROW_YEARISH_TOKENS):
        return True
    # month + year in same string
    if any(m in s for m in MONTH_TOKENS) and any(rx.search(s) for rx in YEAR_REGEXES):
        return True
    return False

def _is_four_digit_year_num(v) -> bool:
    if isinstance(v, (int, float)) and not isinstance(v, bool) and float(v).is_integer():
        iv = int(v)
        return 1900 <= iv <= 2100
    return False

def _collect_yearish_columns(ws: Worksheet, scan_rows: int = 6) -> Set[int]:
    """
    Look at the first few rows. If a cell in a column contains a year-ish
    string/value, mark the whole column 'yearish'.
    """
    yearish_cols: Set[int] = set()
    max_scan_row = min(ws.max_row or 0, scan_rows)
    for r in range(1, max_scan_row + 1):
        for row in ws.iter_rows(min_row=r, max_row=r, values_only=False):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and _cell_has_yearish_text(v):
                    yearish_cols.add(cell.column)
                elif _is_four_digit_year_num(v):
                    # Numbers in header area often represent years
                    yearish_cols.add(cell.column)
    return yearish_cols

def _row_looks_yearish(row) -> bool:
    texts = []
    for c in row:
        v = c.value
        if isinstance(v, str) and v:
            texts.append(v)
    if not texts:
        return False
    row_text = normalize_header(" ".join(map(str, texts)))
    return _cell_has_yearish_text(row_text)

def _looks_like_year(cell, hdr_text: str, v, column_yearish: bool, row_yearish: bool) -> bool:
    # 1) Real datetime
    if isinstance(v, datetime):
        return True

    # 2) Integer in 1900–2100 with year-ish context
    if _is_four_digit_year_num(v):
        if column_yearish or row_yearish:
            return True
        fmt = (cell.number_format or "").lower()
        header_yearish = any(tok in hdr_text for tok in YEARISH_HEADER_TOKENS)
        fmt_yearish = any(x in fmt for x in ["yy", "yyyy", "mmm", "mm/", "dd", "-yy", "d-m", "m-d"])
        if header_yearish or fmt_yearish:
            return True

    return False

# ---------------- Main processing ----------------

def process_excel(
    input_bytes: bytes,
    mode: str = "thousand",                  # thousand | lakh | auto
    header_row: int = 1,
    lakh_edge_threshold: float = 50000,      # < threshold → thousand fallback in lakh/auto
    sheets_include: Optional[List[str]] = None,
    sheets_exclude: Optional[List[str]] = None,
) -> Tuple[bytes, dict]:
    """
    Process an Excel file given as bytes and return (output_bytes, summary_dict).
    Automatically rounds 'No of days' on Depreciation sheets (behavior unchanged).
    """
    import io

    # Load workbook (evaluated values only)
    bio = io.BytesIO(input_bytes)
    wb = load_workbook(bio, data_only=True)

    # Filter sheets if requested
    def _sheet_allowed(name: str) -> bool:
        allowed = True
        if sheets_include:
            allowed = any(pat.lower() in name.lower() for pat in sheets_include)
        if sheets_exclude and any(pat.lower() in name.lower() for pat in sheets_exclude):
            allowed = False
        return allowed

    summary = {
        "mode": mode,
        "header_row": header_row,
        "lakh_edge_threshold": lakh_edge_threshold,
        "sheets": {},
        "totals": {"cells_seen": 0, "cells_converted": 0},
    }

    for ws in wb.worksheets:
        if not _sheet_allowed(ws.title):
            continue

        title = (ws.title or "").strip()
        title_l = title.lower()
        is_depr = "depreciation" in title_l
        hdrs = header_map(ws, header_row) if ws.max_row >= header_row else {}

        # NEW: detect columns that look like years in the first few rows
        yearish_cols = _collect_yearish_columns(ws, scan_rows=6)

        days_cols: Set[int] = set()
        if is_depr and hdrs:
            for c, h in hdrs.items():
                if (h in DAYS_HEADER_VARIANTS) or ("day" in h and "holiday" not in h):
                    days_cols.add(c)

        ws_stats = {"cells_seen": 0, "cells_converted": 0}

        for row in ws.iter_rows():
            row_yearish = _row_looks_yearish(row)  # NEW: row context
            for cell in row:
                # Skip non-master merged cells
                if isinstance(cell, MergedCell):
                    continue

                v = cell.value
                ws_stats["cells_seen"] += 1

                # Depreciation "No of days" rounding (as in your original; left non-destructive)
                if is_depr and days_cols and cell.column in days_cols and isinstance(v, (int, float)) and not isinstance(v, bool):
                    _ = round_half_up_int(float(v))
                    # If you want to actually write it back, uncomment:
                    # if _ != v:
                    #     cell.value = _
                    #     ws_stats["cells_converted"] += 1
                    continue

                # Amount conversion for significant numerics
                if isinstance(v, (int, float, datetime)) and not isinstance(v, bool):
                    # Skip percentages
                    if is_percentage_cell(cell):
                        continue

                    hdr_text = hdrs.get(cell.column, "")
                    col_yearish = (cell.column in yearish_cols)

                    # Skip if this looks like a year/date field
                    if _looks_like_year(cell, hdr_text, v, col_yearish, row_yearish):
                        continue

                    # Only treat larger numbers as amounts
                    if isinstance(v, (int, float)) and abs(float(v)) >= 100:
                        new_v = _divide_amount(float(v), mode, lakh_edge_threshold)
                        if new_v != v:
                            cell.value = new_v
                            ws_stats["cells_converted"] += 1

        summary["sheets"][title] = ws_stats
        summary["totals"]["cells_seen"] += ws_stats["cells_seen"]
        summary["totals"]["cells_converted"] += ws_stats["cells_converted"]

    # Save to bytes
    out_bio = io.BytesIO()
    wb.save(out_bio)
    out_bio.seek(0)

    return out_bio.read(), summary


# -------- Helpers for saving outputs/logs on server/UI ----------

def save_outputs(
    output_bytes: bytes,
    summary: dict,
    original_filename: str,
    out_dir: str = "backup",
    tag: str = "converted",
) -> Tuple[str, str]:
    os.makedirs(out_dir, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.splitext(os.path.basename(original_filename))[0]
    xlsx_path = os.path.join(out_dir, f"{base}__{tag}__{stamp}.xlsx")
    json_path = os.path.join(out_dir, f"{base}__{tag}__{stamp}.json")

    with open(xlsx_path, "wb") as f:
        f.write(output_bytes)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    return xlsx_path, json_path
